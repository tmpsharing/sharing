import openpyxl
from ruamel.yaml import YAML
from ruamel.yaml.comments import CommentedMap
#
yaml = YAML()
yaml.preserve_quotes = True
yaml.explicit_start = True
# # Чтение YAML файла
with open('data4.yaml', 'r', encoding='utf-8') as file:
    data = yaml.load(file)
#
# # Предположим, что obj1 - это словарь и мы хотим его отредактировать
# obj1 = data.get('obj1')
#
# # Пример изменения obj1
# if obj1 is not None:
#     obj1['new_key'] = 'new_value'
#     data['obj1'] = obj1
#
# # Сохранение изменений обратно в YAML файл
# with open('temp.yaml', 'w', encoding='utf-8') as file:
#     yaml.dump(data, file)


def read_excel_sheet(sheet_name, workbook):
    sheet = workbook[sheet_name]
    data_dict = {}

    # Iterate through rows 2 to 10 (assuming data starts from row 2)
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if row[0] is None:
            break
        key, value = row
        data_dict[key] = value

    return data_dict


# Load the Excel workbook
workbook = openpyxl.load_workbook('example1.xlsx')
DEFAULT_SHEET_NAME = "default"
# Get all sheet names
sheet_names = workbook.sheetnames

idms_sheet_names = sheet_names.copy()
idms_sheet_names.remove(DEFAULT_SHEET_NAME)

defaultMapping = read_excel_sheet(DEFAULT_SHEET_NAME, workbook)
overrides = list(map(lambda x: {x[-3:]: read_excel_sheet(x, workbook)}, idms_sheet_names))

# Print all sheet names
# for name in sheet_names:
#     print(read_excel_sheet(name, workbook))

# with open('data4.yaml', 'r', encoding='utf-8') as file:
#     data = yaml.safe_load(file)

# idms_mapping = data.get('idms_mapping', CommentedMap())
#
# CommentedMap()
#
# data['idms_mapping'] = {
#     DEFAULT_SHEET_NAME: defaultMapping,
#     "overrides": overrides
# }


# result = {
#     "idms_mapping" : {
#         DEFAULT_SHEET_NAME: defaultMapping,
#         "overrides": overrides
#     }
# }

with open('temp.yaml', 'w', encoding='utf-8') as file:
    yaml.dump(data, file)


print(overrides)
