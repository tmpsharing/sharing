import openpyxl
import yaml

def read_excel_sheet(sheet_name, workbook):
    sheet = workbook[sheet_name]
    data_dict = {}

    # Iterate through rows 2 to 10 (assuming data starts from row 2)
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if row[0] is None:
            break
        key, value = row
        data_dict[key] = value

    return data_dict


# Load the Excel workbook
workbook = openpyxl.load_workbook('example1.xlsx')
DEFAULT_SHEET_NAME = "default"
# Get all sheet names
sheet_names = workbook.sheetnames

idms_sheet_names = sheet_names.copy()
idms_sheet_names.remove(DEFAULT_SHEET_NAME)

defaultMapping = read_excel_sheet(DEFAULT_SHEET_NAME, workbook)
overrides = list(map(lambda x: {x[-3:]: read_excel_sheet(x, workbook)}, idms_sheet_names))

# Print all sheet names
# for name in sheet_names:
#     print(read_excel_sheet(name, workbook))

with open('data4.yaml', 'r', encoding='utf-8') as file:
    data = yaml.safe_load(file)

data['idms_mapping'] = {
    DEFAULT_SHEET_NAME: defaultMapping,
    "overrides": overrides
}


# result = {
#     "idms_mapping" : {
#         DEFAULT_SHEET_NAME: defaultMapping,
#         "overrides": overrides
#     }
# }
with open('temp.yaml', 'w', encoding='utf-8') as file:
    yaml.dump(data, file, default_flow_style=False, allow_unicode=True)

# with open('data4.yaml', 'w') as yaml_file:
#     yaml.dump(result, yaml_file, default_flow_style=False)

print(overrides)
